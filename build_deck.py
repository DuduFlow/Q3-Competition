from __future__ import annotations

import html
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
TEMPLATE = Path("/Users/paulbastar/Downloads/deck-base.html")
WORKBOOK = Path("/Users/paulbastar/Downloads/複本 115年第三季競賽報表_0806.xlsx")
OUTPUT = ROOT / "index.html"
XLSX_LIB = ROOT / "xlsx.full.min.js"
EXCEL_REFRESH = ROOT / "excel-refresh.js"

REPORT_TITLE = "115年第三季競賽報表"
REPORT_TITLE_FULL = "115年第三季競賽報表｜全體戰情"
REPORT_SIG = "115 Q3 Contest"
REPORT_PERIOD = "競賽期間 115/06/26-115/09/29 · 統計至 115/08/06"
MONTH_LABELS = ("七月", "八月", "九月")

TARGET_OFFICES = ["TP567", "TP512", "TP591", "TP767"]
GROUPS = ["個人組", "新人組"]

REACH_RULES = {
    "個人組": {
        "monthly": 50000,
        "quarter": 180000,
        "reward": 25000,
        "label": "每月5萬或季累計18萬",
    },
    "新人組": {
        "monthly": 0,
        "quarter": 80000,
        "reward": 3000,
        "label": "季累計8萬/11萬/14萬階梯獎勵",
    },
}

NEWCOMER_TARGET_TIERS = [
    (180000, 13000, "18萬加發"),
    (140000, 12000, "14萬達標"),
    (110000, 6000, "11萬達標"),
    (80000, 3000, "8萬達標"),
]


COLUMNS = {
    "rank": 0,
    "division": 1,
    "region": 2,
    "office_code": 3,
    "office_name": 4,
    "name": 5,
    "apr_verified": 6,
    "apr_pending": 7,
    "apr_total": 8,
    "may_verified": 9,
    "may_pending": 10,
    "may_total": 11,
    "jun_verified": 12,
    "jun_pending": 13,
    "jun_total": 14,
    "personal_verified": 15,
    "personal_pending": 16,
    "personal_total": 17,
    "critical_bonus": 18,
    "care_bonus": 19,
    "medical_bonus": 20,
    "participating_bonus": 21,
    "interest_bonus": 22,
    "buyback_bonus": 23,
    "group_bonus": 24,
    "contest_verified": 25,
    "contest_pending": 26,
    "contest_total": 27,
    "personal_no_investment": 28,
    "base_period": 29,
}


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"[\ue000-\uf8ff]", "", text)


def num(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    try:
        return int(round(float(str(value).replace(",", ""))))
    except ValueError:
        return 0


def record_from_row(group: str, row: tuple) -> dict:
    rec = {"group": group}
    for key, idx in COLUMNS.items():
        value = row[idx] if idx < len(row) else None
        if key in {"division", "region", "office_code", "office_name", "name", "base_period"}:
            rec[key] = clean_text(value)
        else:
            rec[key] = num(value)
    rec.update(reward_progress(rec))
    return rec


def over_reward(group: str, contest_fyc: int, qualified: bool) -> int:
    if not qualified:
        return 0
    if group == "個人組":
        if contest_fyc < 250000:
            return 0
        return 10000 + max(0, (contest_fyc - 250000) // 100000) * 15000
    if contest_fyc < 200000:
        return 0
    if contest_fyc < 250000:
        return 10000
    return 20000 + max(0, (contest_fyc - 250000) // 100000) * 15000


def next_over_gap(group: str, contest_fyc: int, qualified: bool) -> tuple[str, int]:
    if not qualified:
        return "先取得達標", 0
    if group == "個人組":
        if contest_fyc < 250000:
            return "25萬", 250000 - contest_fyc
        next_target = 250000 + ((contest_fyc - 250000) // 100000 + 1) * 100000
        return f"{next_target // 10000}萬", next_target - contest_fyc
    if contest_fyc < 200000:
        return "20萬", 200000 - contest_fyc
    if contest_fyc < 250000:
        return "25萬", 250000 - contest_fyc
    next_target = 250000 + ((contest_fyc - 250000) // 100000 + 1) * 100000
    return f"{next_target // 10000}萬", next_target - contest_fyc


def reward_progress(rec: dict) -> dict:
    rule = REACH_RULES[rec["group"]]
    if rec["group"] == "個人組":
        monthly_ok = all(rec[key] >= rule["monthly"] for key in ("apr_total", "may_total", "jun_total"))
        quarter_ok = rec["personal_total"] >= rule["quarter"]
        qualified = monthly_ok or quarter_ok
        quarter_gap = max(0, rule["quarter"] - rec["personal_total"])
        monthly_gap = sum(max(0, rule["monthly"] - rec[key]) for key in ("apr_total", "may_total", "jun_total"))
        reach_gap = 0 if qualified else min(quarter_gap, monthly_gap)
        target_reward = rule["reward"] if qualified else 0
        reach_threshold = rule["quarter"]
        reach_note = "季累計達標" if qualified and quarter_ok else ("月月達標" if qualified else f"差{money(reach_gap)}")
    else:
        target_reward = 0
        reach_note = ""
        for threshold, reward, label in NEWCOMER_TARGET_TIERS:
            if rec["personal_total"] >= threshold:
                target_reward = reward
                reach_note = label
                break
        reach_threshold = NEWCOMER_TARGET_TIERS[-1][0]
        qualified = target_reward > 0
        reach_gap = 0 if qualified else max(0, reach_threshold - rec["personal_total"])
        if not qualified:
            reach_note = f"差{money(reach_gap)}"
    over = over_reward(rec["group"], rec["contest_total"], qualified)
    next_label, next_gap = next_over_gap(rec["group"], rec["contest_total"], qualified)
    return {
        "reach_rule": rule["label"],
        "reach_qualified": qualified,
        "reach_note": reach_note,
        "reach_gap": reach_gap,
        "reach_threshold": reach_threshold,
        "target_reward": target_reward,
        "over_reward": over,
        "total_reward": target_reward + over,
        "next_over_label": next_label,
        "next_over_gap": next_gap,
    }


def extract_records() -> list[dict]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    records: list[dict] = []
    for group in GROUPS:
        ws = wb[group]
        for row in ws.iter_rows(values_only=True):
            office_code = clean_text(row[COLUMNS["office_code"]] if len(row) > COLUMNS["office_code"] else "")
            rank = row[COLUMNS["rank"]] if len(row) > COLUMNS["rank"] else None
            if office_code in TARGET_OFFICES and isinstance(rank, (int, float)):
                records.append(record_from_row(group, row))
    return records


def money(value: int) -> str:
    return f"{int(value):,}"


def compact_money(value: int) -> str:
    if value >= 10000:
        return f"{value / 10000:.1f}萬"
    return money(value)


def pct(part: int, total: int) -> str:
    return "0%" if not total else f"{part / total:.1%}"


def esc(value) -> str:
    return html.escape(str(value), quote=True)


def summarize(records: list[dict]) -> dict:
    summary = {}
    for group in GROUPS:
        group_records = [r for r in records if r["group"] == group]
        summary[group] = {
            "count": len(group_records),
            "personal_total": sum(r["personal_total"] for r in group_records),
            "contest_total": sum(r["contest_total"] for r in group_records),
            "reach_count": sum(1 for r in group_records if r["reach_qualified"]),
            "over_count": sum(1 for r in group_records if r["over_reward"] > 0),
            "target_reward": sum(r["target_reward"] for r in group_records),
            "over_reward": sum(r["over_reward"] for r in group_records),
            "total_reward": sum(r["total_reward"] for r in group_records),
            "offices": {},
        }
        for office in TARGET_OFFICES:
            rows = [r for r in group_records if r["office_code"] == office]
            summary[group]["offices"][office] = {
                "count": len(rows),
                "office_name": rows[0]["office_name"] if rows else "",
                "personal_total": sum(r["personal_total"] for r in rows),
                "contest_total": sum(r["contest_total"] for r in rows),
                "reach_count": sum(1 for r in rows if r["reach_qualified"]),
                "over_count": sum(1 for r in rows if r["over_reward"] > 0),
                "target_reward": sum(r["target_reward"] for r in rows),
                "over_reward": sum(r["over_reward"] for r in rows),
                "total_reward": sum(r["total_reward"] for r in rows),
                "top": max(rows, key=lambda r: r["contest_total"]) if rows else None,
            }
    return summary


def top_rows(records: list[dict], group: str, office: str | None = None, limit: int = 12) -> list[dict]:
    rows = [r for r in records if r["group"] == group]
    if office:
        rows = [r for r in rows if r["office_code"] == office]
    return sorted(rows, key=lambda r: (r["contest_total"], r["personal_total"]), reverse=True)[:limit]


def full_rows(records: list[dict], group: str, office: str | None = None) -> list[dict]:
    rows = [r for r in records if r["group"] == group]
    if office:
        rows = [r for r in rows if r["office_code"] == office]
    return sorted(rows, key=lambda r: (r["contest_total"], r["personal_total"]), reverse=True)


def person_table(rows: list[dict], full: bool = False) -> str:
    if not rows:
        return '<div class="empty">沒有資料</div>'
    body = []
    for i, r in enumerate(rows, start=1):
        if full:
            body.append(
                "<tr>"
                f"<td>{i}</td>"
                f"<td><strong>{esc(r['name'])}</strong><span class=\"person-unit\">（{esc(r['office_code'])} {esc(r['office_name'])}）</span></td>"
                f"<td>{money(r['apr_total'])}</td>"
                f"<td>{money(r['may_total'])}</td>"
                f"<td>{money(r['jun_total'])}</td>"
                f"<td>{money(r['personal_total'])}</td>"
                f"<td>{money(r['contest_total'])}</td>"
                f"<td>{status_badge(r)}</td>"
                f"<td>{money(r['over_reward'])}</td>"
                f"<td>{money(r['total_reward'])}</td>"
                "</tr>"
            )
        else:
            body.append(
                "<tr>"
                f"<td>{i}</td>"
                f"<td><strong>{esc(r['name'])}</strong><span class=\"person-unit\">（{esc(r['office_code'])} {esc(r['office_name'])}）</span></td>"
                f"<td>{money(r['personal_total'])}</td>"
                f"<td>{money(r['contest_total'])}</td>"
                f"<td>{status_badge(r)}</td>"
                f"<td>{money(r['over_reward'])}</td>"
                f"<td>{money(r['total_reward'])}</td>"
                "</tr>"
            )
    table_class = "data-table full-table" if full else "data-table"
    if full:
        header = (
            f"<thead><tr><th>#</th><th>姓名（單位）</th><th>{MONTH_LABELS[0]}</th><th>{MONTH_LABELS[1]}</th><th>{MONTH_LABELS[2]}</th>"
            "<th>個人險FYC</th><th>競賽FYC</th><th>達標</th><th>超標獎金</th><th>預估合計</th></tr></thead>"
        )
    else:
        header = (
            "<thead><tr><th>#</th><th>姓名（單位）</th><th>個人險FYC</th><th>競賽FYC</th>"
            "<th>達標</th><th>超標獎金</th><th>預估合計</th></tr></thead>"
        )
    return (
        f'<table class="{table_class}">'
        f"{header}"
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def status_badge(r: dict) -> str:
    if r["reach_qualified"]:
        return f'<span class="status-badge ok">{esc(r["reach_note"])}</span>'
    return f'<span class="status-badge wait">{esc(r["reach_note"])}</span>'


def page_num(n: int, total: int) -> str:
    return f"{n:02d} / {total:02d}"


def sig() -> str:
    return f'<div class="slide-sig"><span class="dot"></span>{REPORT_SIG}</div>'


def pg(n: int, total: int) -> str:
    return f'<div class="slide-pg">{page_num(n, total)}</div>'


def cover(summary: dict, total_pages: int) -> str:
    return f"""
  <section class="slide cover title-slide">
    <div class="kicker animate-line" style="--i:0">全體合併呈現 · 每日 Excel 更新</div>
    <h1 class="animate-line" style="--i:1">{REPORT_TITLE}</h1>
    <p class="animate-line meta" style="--i:2">{REPORT_PERIOD}</p>
    <div class="cover-metrics deck-steps animate-line" style="--i:3">
      <div><span>01</span><label>選擇今日 Excel</label></div>
      <div><span>02</span><label>自動重算獎勵與資格</label></div>
      <div><span>03</span><label>投影佈達競賽進度</label></div>
    </div>
    <div class="upload-panel animate-line" style="--i:4">
      <label for="excel-upload">讀取每日更新 Excel</label>
      <div class="upload-row">
        <input id="excel-upload" type="file" accept=".xlsx,.xls,.xlsm">
        <span id="excel-upload-status">目前使用內建資料，可選擇新版 Excel 重新計算。</span>
      </div>
    </div>
    {sig()}
    {pg(1, total_pages)}
  </section>
"""


def overview(summary: dict, total_pages: int) -> str:
    personal = summary["個人組"]
    newcomer = summary["新人組"]
    total_count = personal["count"] + newcomer["count"]
    total_personal = personal["personal_total"] + newcomer["personal_total"]
    total_contest = personal["contest_total"] + newcomer["contest_total"]
    total_reach = personal["reach_count"] + newcomer["reach_count"]
    total_over = personal["over_count"] + newcomer["over_count"]
    total_reward = personal["total_reward"] + newcomer["total_reward"]
    not_reached = max(total_count - total_reach, 0)
    return f"""
  <section class="slide slide-grid">
    <div class="slide-header">
      <div><p class="meta">OVERVIEW</p><h2>全體戰情總覽</h2></div>
      <div class="pill">Excel 擷取版</div>
    </div>
    <div class="hero-metrics">
      <div><span>{total_count}</span><label>參賽筆數</label></div>
      <div><span>{compact_money(total_contest)}</span><label>累計競賽FYC</label></div>
      <div><span>{compact_money(total_reward)}</span><label>目前預估獎金</label></div>
    </div>
    <div class="summary-strip">
      <div><span>{total_reach}</span><label>已達標</label></div>
      <div><span>{total_over}</span><label>已有超標獎金</label></div>
      <div><span>{not_reached}</span><label>尚未達標</label></div>
      <div><span>{compact_money(total_personal)}</span><label>個人險FYC</label></div>
    </div>
    <div class="progress-cards">
      <article class="progress-card">
        <div class="progress-title">個人組</div>
        <div class="progress-big">{personal['reach_count']} / {personal['count']}</div>
        <div class="progress-sub">已達標 · 競賽FYC {compact_money(personal['contest_total'])}</div>
        <div class="bar"><span style="width:{personal['reach_count'] / personal['count'] * 100 if personal['count'] else 0:.1f}%"></span></div>
      </article>
      <article class="progress-card">
        <div class="progress-title">新人組</div>
        <div class="progress-big">{newcomer['reach_count']} / {newcomer['count']}</div>
        <div class="progress-sub">已達標 · 競賽FYC {compact_money(newcomer['contest_total'])}</div>
        <div class="bar"><span style="width:{newcomer['reach_count'] / newcomer['count'] * 100 if newcomer['count'] else 0:.1f}%"></span></div>
      </article>
    </div>
    {sig()}
    {pg(2, total_pages)}
  </section>
"""


def rules_slide(total_pages: int) -> str:
    return f"""
  <section class="slide rule-slide">
    <div class="slide-header">
      <div><p class="meta">RULES</p><h2>本版呈現：Excel 達標與超標進度</h2></div>
      <div class="pill">以 Excel 數字為主</div>
    </div>
    <div class="rule-grid">
      <article class="rule-card primary">
        <p class="meta">達標獎勵</p>
        <h3>只看個人險 FYC</h3>
        <div class="rule-amount">個人組 25,000</div>
        <ul>
          <li>個人組：每月皆達 5 萬，或季累計達 18 萬。</li>
          <li>新人組：季累計 8 萬 / 11 萬 / 14 萬，獎金 3,000 / 6,000 / 12,000。</li>
          <li>新人組個人險 FYC 達 18 萬，另加發 1,000 元。</li>
          <li>加碼商品、團險、房貸等加成，不拿來判斷達標。</li>
        </ul>
      </article>
      <article class="rule-card accent">
        <p class="meta">超標獎勵</p>
        <h3>先達標，再看競賽 FYC</h3>
        <div class="rule-amount">無上限</div>
        <ul>
          <li>個人組：競賽 FYC 達 25 萬，超標獎金 10,000 元。</li>
          <li>新人組：競賽 FYC 達 20 萬，超標獎金 10,000 元；達 25 萬為 20,000 元。</li>
          <li>25 萬之後，每增加 10 萬再加 15,000 元。</li>
        </ul>
      </article>
    </div>
    <div class="rule-note">預估合計只計入達標與超標獎勵；高標需另看個人險 FYC 成長率與基期，本版先作規則提醒，不混入獎金合計。</div>
    {sig()}
    {pg(3, total_pages)}
  </section>
"""


def travel_slide(records: list[dict], summary: dict, total_pages: int) -> str:
    rows = sorted(
        [r for r in records if r["group"] == "新人組"],
        key=lambda r: (r["personal_total"], r["contest_total"]),
        reverse=True,
    )[:10]
    total_reached = summary["新人組"]["reach_count"]
    body = []
    for i, r in enumerate(rows, start=1):
        badge_class = "ok" if r["reach_qualified"] else "wait"
        next_label = newcomer_next_tier(r["personal_total"])
        body.append(
            "<tr>"
            f"<td>{i}</td>"
            f"<td><strong>{esc(r['name'])}</strong><span class=\"person-unit\">（{esc(r['office_code'])} {esc(r['office_name'])} · {esc(r['group'])}）</span></td>"
            f"<td>{money(r['personal_total'])}</td>"
            f"<td>{money(r['target_reward'])}</td>"
            f"<td><span class=\"status-badge {badge_class}\">{esc(r['reach_note'])}</span></td>"
            f"<td>{esc(next_label)}</td>"
            "</tr>"
        )
    return f"""
  <section class="slide travel-slide focus-slide">
    <div class="slide-header">
      <div><p class="meta">NEWCOMER TARGET</p><h2>新人組達標階梯 · 個人險 FYC</h2></div>
      <div class="pill success">已達標 {total_reached} 人</div>
    </div>
    <table class="data-table focus-table">
      <thead><tr><th>#</th><th>姓名（單位）</th><th>個人險FYC</th><th>達標獎金</th><th>目前級距</th><th>下一段</th></tr></thead>
      <tbody>{''.join(body)}</tbody>
    </table>
    <div class="rule-note travel-note">新人組第三季達標獎勵：8 萬 3,000、11 萬 6,000、14 萬 12,000；達 18 萬另加 1,000。</div>
    {sig()}
    {pg(4, total_pages)}
  </section>
"""


def newcomer_next_tier(personal_total: int) -> str:
    for threshold, _reward, label in reversed(NEWCOMER_TARGET_TIERS):
        if personal_total < threshold:
            return f"{label}差 {money(threshold - personal_total)}"
    return "最高級距"


def reward_summary_slide(summary: dict, total_pages: int, page: int = 5) -> str:
    personal = summary["個人組"]
    newcomer = summary["新人組"]
    total_reach = personal["reach_count"] + newcomer["reach_count"]
    total_over = personal["over_count"] + newcomer["over_count"]
    total_reward = personal["total_reward"] + newcomer["total_reward"]
    cards = []
    for group, data in summary.items():
        cards.append(
            f"""
      <article class="progress-card">
        <div class="progress-title">{group}</div>
        <div class="progress-big">{data['reach_count']} / {data['count']}</div>
        <div class="progress-sub">已取得達標獎勵</div>
        <div class="bar"><span style="width:{data['reach_count'] / data['count'] * 100 if data['count'] else 0:.1f}%"></span></div>
        <div class="progress-stats">
          <div><b>{data['over_count']}</b><span>超標</span></div>
          <div><b>{compact_money(data['target_reward'])}</b><span>達標獎金</span></div>
          <div><b>{compact_money(data['over_reward'])}</b><span>超標獎金</span></div>
        </div>
      </article>
"""
        )
    return f"""
  <section class="slide reward-overview">
    <div class="slide-header">
      <div><p class="meta">REWARD PROGRESS</p><h2>獎勵進度總覽</h2></div>
      <div class="pill success">預估獎金 {compact_money(total_reward)}</div>
    </div>
    <div class="hero-metrics">
      <div><span>{total_reach}</span><label>已達標人數</label></div>
      <div><span>{total_over}</span><label>已進入超標獎勵</label></div>
      <div><span>{compact_money(total_reward)}</span><label>達標＋超標預估獎金</label></div>
    </div>
    <div class="progress-cards">{''.join(cards)}</div>
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def reach_progress_slide(records: list[dict], total_pages: int, page: int = 6) -> str:
    almost = sorted(
        [r for r in records if not r["reach_qualified"]],
        key=lambda r: (r["reach_gap"], -r["personal_total"]),
    )[:10]
    reached_count = sum(1 for r in records if r["reach_qualified"])
    return f"""
  <section class="slide reach-slide focus-slide">
    <div class="slide-header">
      <div><p class="meta">REACH TARGET</p><h2>達標衝刺 · 最接近達標 Top 10</h2></div>
      <div class="pill">已達標 {reached_count} 人</div>
    </div>
    {focus_reward_table(almost, mode="reach")}
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def over_progress_slide(records: list[dict], total_pages: int, page: int = 7) -> str:
    next_rows = sorted(
        [r for r in records if r["reach_qualified"]],
        key=lambda r: (r["next_over_gap"], -r["contest_total"]),
    )[:10]
    over_count = sum(1 for r in records if r["over_reward"] > 0)
    return f"""
  <section class="slide over-slide focus-slide">
    <div class="slide-header">
      <div><p class="meta">OVER TARGET</p><h2>超標衝刺 · 下一段門檻 Top 10</h2></div>
      <div class="pill orange">已有超標 {over_count} 人</div>
    </div>
    {focus_reward_table(next_rows, mode="over")}
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def focus_reward_table(rows: list[dict], mode: str) -> str:
    if not rows:
        return '<div class="empty">沒有資料</div>'
    body = []
    for i, r in enumerate(rows, start=1):
        if mode == "over":
            status = f"{r['next_over_label']} 差 {money(r['next_over_gap'])}"
            extra_label = "目前超標獎金"
            extra_value = money(r["over_reward"])
        else:
            status = f"達標差 {money(r['reach_gap'])}"
            extra_label = "達標門檻"
            extra_value = r.get("reach_threshold", REACH_RULES[r["group"]]["quarter"] if r["group"] in REACH_RULES else 0)
            extra_value = money(extra_value)
        body.append(
            "<tr>"
            f"<td>{i}</td>"
            f"<td><strong>{esc(r['name'])}</strong><span class=\"person-unit\">（{esc(r['office_code'])} {esc(r['office_name'])} · {esc(r['group'])}）</span></td>"
            f"<td>{money(r['personal_total'])}</td>"
            f"<td>{money(r['contest_total'])}</td>"
            f"<td>{extra_value}</td>"
            f"<td><span class=\"status-badge wait\">{esc(status)}</span></td>"
            "</tr>"
        )
    return (
        '<table class="data-table focus-table">'
        f"<thead><tr><th>#</th><th>姓名（單位）</th><th>個人險FYC</th><th>競賽FYC</th><th>{extra_label}</th><th>差距</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def mini_reward_table(rows: list[dict], show_gap: bool = False, show_reward: bool = False, show_next: bool = False) -> str:
    if not rows:
        return '<div class="empty">沒有資料</div>'
    body = []
    for i, r in enumerate(rows, start=1):
        if show_reward:
            note = f"超標 {money(r['over_reward'])}"
        elif show_next:
            note = f"{r['next_over_label']} 差 {money(r['next_over_gap'])}"
        elif show_gap:
            note = f"達標差 {money(r['reach_gap'])}"
        else:
            note = r["reach_note"]
        body.append(
            "<tr>"
            f"<td>{i}</td>"
            f"<td><strong>{esc(r['name'])}</strong><span class=\"person-unit\">（{esc(r['office_code'])} {esc(r['office_name'])} · {esc(r['group'])}）</span></td>"
            f"<td>{money(r['personal_total'])}</td>"
            f"<td>{money(r['contest_total'])}</td>"
            f"<td>{esc(note)}</td>"
            "</tr>"
        )
    return (
        '<table class="mini-table">'
        "<thead><tr><th>#</th><th>姓名</th><th>FYC</th><th>競賽FYC</th><th>狀態</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def ranking_slide(records: list[dict], group: str, page: int, total_pages: int) -> str:
    rows = top_rows(records, group, limit=10)
    return f"""
  <section class="slide ranking-slide">
    <div class="slide-header">
      <div><p class="meta">RANKING</p><h2>{group} · 累計競賽FYC Top 10</h2></div>
      <div class="pill orange">{len(full_rows(records, group))} 筆</div>
    </div>
    {person_table(rows)}
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def office_slide(records: list[dict], summary: dict, office: str, page: int, total_pages: int) -> str:
    personal = summary["個人組"]["offices"][office]
    newcomer = summary["新人組"]["offices"][office]
    office_name = personal["office_name"] or newcomer["office_name"]
    personal_rows = top_rows(records, "個人組", office, limit=6)
    newcomer_rows = top_rows(records, "新人組", office, limit=6)
    return f"""
  <section class="slide detail-slide">
    <div class="slide-header">
      <div><p class="meta">OFFICE DETAIL</p><h2>{office} {esc(office_name)}</h2></div>
      <div class="pill success">個人 {personal['count']} · 新人 {newcomer['count']}</div>
    </div>
    <div class="detail-metrics">
      <div><span>{personal['reach_count'] + newcomer['reach_count']}</span><label>已達標人數</label></div>
      <div><span>{personal['over_count'] + newcomer['over_count']}</span><label>超標獎金人數</label></div>
      <div><span>{compact_money(personal['total_reward'] + newcomer['total_reward'])}</span><label>預估獎金合計</label></div>
    </div>
    <div class="two-col">
      <div>
        <h3>個人組前 6 名</h3>
        {person_table(personal_rows)}
      </div>
      <div>
        <h3>新人組前 6 名</h3>
        {person_table(newcomer_rows)}
      </div>
    </div>
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def full_data_slide(records: list[dict], group: str, page: int, total_pages: int) -> str:
    rows = full_rows(records, group)
    return f"""
  <section class="slide full-data">
    <div class="slide-header">
      <div><p class="meta">FULL DATA</p><h2>{group} · 全體完整清單</h2></div>
      <div class="header-actions">
        <button class="pdf-export-btn" type="button" onclick="exportFullListPdf('{group}')">輸出PDF</button>
        <div class="pill gold">{len(rows)} 筆資料</div>
      </div>
    </div>
    <div class="table-scroll">
      {person_table(rows, full=True)}
    </div>
    {sig()}
    {pg(page, total_pages)}
  </section>
"""


def custom_css() -> str:
    return """
:root {
  --brand: #0093C1;
  --accent: #FFC94F;
  --orange: #FF8933;
  --success: #30DAA2;
  --gold: #FFC94F;
  --danger: #E03E57;
  --surface: #F8FAFC;
  --surface-dark: #101820;
  --text: #23303D;
  --text-strong: #0B1720;
  --text-muted: #68778A;
  --text-on-dark: #F6FAFD;
  --font-sans: "Noto Sans TC", "PingFang TC", "Microsoft JhengHei", Arial, sans-serif;
  --font-display: "Noto Sans TC", "PingFang TC", "Microsoft JhengHei", Arial, sans-serif;
  --fs-cover: 5.8em;
  --fs-heading: 3.25em;
  --fs-subtitle: 1.6em;
  --fs-body: 1.12em;
  --fs-meta: 0.88em;
  --lh-tight: 1.08;
  --lh-body: 1.5;
}

.slide:not(.cover) {
  justify-content: flex-start;
  padding-top: 5.4em;
}

.title-slide {
  background:
    linear-gradient(130deg, rgba(0,147,193,0.92), rgba(16,24,32,0.92)),
    radial-gradient(circle at 75% 20%, rgba(255,201,79,0.38), transparent 30%),
    #101820;
}

.kicker {
  color: var(--accent);
  font-size: 1.1em;
  font-weight: 800;
  letter-spacing: 0.14em;
  margin-bottom: 0.9em;
}

.cover-metrics {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 1.1em;
  width: min(980px, 78%);
  margin-top: 3.2em;
}

.upload-panel {
  width: min(980px, 78%);
  margin-top: 1.4em;
  padding: 1em 1.2em;
  border-radius: 8px;
  background: rgba(255,255,255,0.12);
  border: 1px solid rgba(255,255,255,0.18);
}

.upload-panel label {
  display: block;
  color: var(--accent);
  font-size: 0.9em;
  font-weight: 900;
  letter-spacing: 0.08em;
  margin-bottom: 0.6em;
}

.upload-row {
  display: flex;
  align-items: center;
  gap: 1em;
}

.upload-row input {
  width: 20em;
  color: rgba(255,255,255,0.85);
  font: inherit;
  font-size: 0.9em;
}

.upload-row input::file-selector-button {
  margin-right: 0.8em;
  border: 0;
  border-radius: 7px;
  padding: 0.55em 0.95em;
  background: var(--accent);
  color: #102332;
  font-weight: 900;
  cursor: pointer;
}

.upload-row span {
  color: rgba(255,255,255,0.76);
  font-size: 0.9em;
  font-weight: 700;
}

.cover-metrics div,
.summary-strip div,
.detail-metrics div {
  background: rgba(255,255,255,0.12);
  border: 1px solid rgba(255,255,255,0.14);
  border-radius: 8px;
  padding: 1.3em 1.5em;
}

.cover-metrics span,
.summary-strip span,
.detail-metrics span {
  display: block;
  font-size: 2.25em;
  font-weight: 900;
  line-height: 1;
}

.cover-metrics label,
.summary-strip label,
.detail-metrics label,
.office-sub,
.office-grid span {
  display: block;
  margin-top: 0.5em;
  color: rgba(255,255,255,0.72);
  font-size: 0.82em;
  font-weight: 700;
}

.slide-header {
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  gap: 2em;
  min-height: 8.7em;
  margin-bottom: 0.9em;
}

.slide-header h2 {
  font-size: 3.6em;
}

.focus-slide .slide-header h2,
.ranking-slide .slide-header h2 {
  font-size: 3.95em;
}

.summary-strip {
  display: grid;
  grid-template-columns: repeat(4, minmax(0, 1fr));
  gap: 1em;
  margin-bottom: 1.5em;
}

.summary-strip div,
.detail-metrics div {
  background: #102332;
  color: #fff;
}

.office-cards {
  display: grid;
  grid-template-columns: repeat(4, minmax(0, 1fr));
  gap: 1em;
}

.office-card {
  background: #fff;
  border: 1px solid #D9E2EC;
  border-radius: 8px;
  padding: 1.35em;
  min-height: 23em;
  box-shadow: 0 18px 50px rgba(24,39,56,0.08);
}

.office-head {
  display: flex;
  align-items: baseline;
  justify-content: space-between;
  gap: 1em;
  border-bottom: 1px solid #E6EDF4;
  padding-bottom: 1em;
  margin-bottom: 1.5em;
}

.office-head span {
  color: var(--brand);
  font-size: 1.35em;
  font-weight: 900;
}

.office-head strong {
  font-size: 1.1em;
  color: var(--text-strong);
}

.office-main {
  font-size: 3em;
  font-weight: 900;
  color: var(--text-strong);
}

.office-card .office-sub,
.office-card .office-grid span {
  color: var(--text-muted);
}

.office-grid {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 0.75em;
  margin-top: 2em;
}

.office-grid div {
  background: #F1F5F9;
  border-radius: 7px;
  padding: 0.9em;
  min-height: 5.2em;
}

.office-grid b {
  display: block;
  color: var(--text-strong);
  font-size: 1.05em;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}

.data-table {
  width: 100%;
  border-collapse: collapse;
  table-layout: fixed;
  background: #fff;
  border: 1px solid #DDE6EF;
  border-radius: 8px;
  overflow: hidden;
  font-size: 1.18em;
}

.data-table th {
  background: #102332;
  color: #fff;
  font-size: 0.84em;
  line-height: 1.25;
  text-align: right;
  padding: 0.85em 0.8em;
  white-space: normal;
}

.data-table th:nth-child(1),
.data-table th:nth-child(2) {
  text-align: left;
}

.data-table td {
  border-top: 1px solid #E7EDF3;
  padding: 0.78em 0.8em;
  text-align: right;
  color: var(--text);
  white-space: nowrap;
}

.data-table td:nth-child(1) {
  width: 3em;
  text-align: left;
  color: var(--text-muted);
}

.data-table td:nth-child(2),
.data-table th:nth-child(2) {
  width: 16em;
  text-align: left;
}

.data-table td strong {
  display: inline;
  color: var(--text-strong);
  font-size: 1.18em;
}

.data-table .person-unit {
  display: inline;
  margin-left: 0.35em;
  color: var(--text-muted);
  font-size: 0.82em;
  font-weight: 700;
}

.data-table td:nth-child(4),
.data-table td:nth-child(7) {
  font-weight: 900;
  color: #0E7698;
}

.focus-slide .data-table,
.ranking-slide .data-table {
  font-size: 1.34em;
}

.focus-slide .data-table th,
.ranking-slide .data-table th {
  font-size: 0.9em;
  padding: 0.9em 0.9em;
}

.focus-slide .data-table td,
.ranking-slide .data-table td {
  padding: 1.04em 0.9em;
}

.focus-slide .data-table td strong,
.ranking-slide .data-table td strong {
  font-size: 1.24em;
}

.focus-slide .data-table .person-unit,
.ranking-slide .data-table .person-unit {
  font-size: 0.86em;
}

.focus-slide .status-badge,
.ranking-slide .status-badge {
  font-size: 1em;
  min-width: 7em;
  padding: 0.38em 0.9em;
}

.focus-slide .data-table td:nth-child(2),
.focus-slide .data-table th:nth-child(2),
.ranking-slide .data-table td:nth-child(2),
.ranking-slide .data-table th:nth-child(2) {
  width: 18em;
}

.focus-slide,
.ranking-slide {
  padding-top: 4.85em;
  padding-bottom: 2.1em;
}

.focus-slide .slide-header,
.ranking-slide .slide-header {
  min-height: 7.65em;
  margin-bottom: 0.65em;
}

.focus-slide .data-table th,
.ranking-slide .data-table th {
  padding-top: 0.76em;
  padding-bottom: 0.76em;
}

.focus-slide .data-table td,
.ranking-slide .data-table td {
  padding-top: 0.82em;
  padding-bottom: 0.82em;
}

.focus-slide .slide-sig,
.ranking-slide .slide-sig {
  bottom: 1.05em;
}

.focus-slide .slide-pg,
.ranking-slide .slide-pg {
  bottom: 1.05em;
}

.two-col {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 1.2em;
  min-height: 0;
}

.two-col h3 {
  margin-bottom: 0.65em;
  font-size: 1.65em;
}

.two-col .data-table {
  font-size: 0.84em;
}

.two-col .data-table td,
.two-col .data-table th {
  padding-left: 0.56em;
  padding-right: 0.56em;
}

.two-col .data-table td:nth-child(2),
.two-col .data-table th:nth-child(2) {
  width: 10.5em;
}

.detail-metrics {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 1em;
  margin-bottom: 1.4em;
}

.hero-metrics {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 1em;
  margin-bottom: 1.4em;
}

.hero-metrics div {
  background: #102332;
  color: #fff;
  border-radius: 8px;
  padding: 1.5em;
}

.hero-metrics span {
  display: block;
  font-size: 3em;
  font-weight: 900;
  line-height: 1;
}

.hero-metrics label {
  display: block;
  margin-top: 0.55em;
  color: rgba(255,255,255,0.72);
  font-size: 0.86em;
  font-weight: 700;
}

.rule-grid,
.progress-cards {
  display: grid;
  grid-template-columns: repeat(2, minmax(0, 1fr));
  gap: 1.2em;
}

.rule-card,
.progress-card {
  background: #fff;
  border: 1px solid #D9E2EC;
  border-radius: 8px;
  padding: 1.65em;
  box-shadow: 0 18px 50px rgba(24,39,56,0.08);
}

.rule-card.primary {
  border-top: 7px solid var(--brand);
}

.rule-card.accent {
  border-top: 7px solid var(--orange);
}

.rule-card h3 {
  font-size: 2em;
  margin: 0.25em 0 0.6em;
}

.rule-card ul {
  display: grid;
  gap: 0.7em;
  padding-left: 1.1em;
}

.rule-card li {
  font-size: 1.16em;
}

.rule-amount {
  display: inline-flex;
  align-items: center;
  min-height: 2.2em;
  margin-bottom: 1.2em;
  padding: 0.35em 0.7em;
  border-radius: 8px;
  background: #F1F5F9;
  color: var(--text-strong);
  font-size: 2em;
  font-weight: 900;
}

.rule-note {
  margin-top: 1.2em;
  padding: 1em 1.2em;
  border-left: 5px solid var(--gold);
  background: #FFF8E1;
  color: #604700;
  border-radius: 8px;
  font-size: 1.08em;
  font-weight: 700;
}

.progress-title {
  color: var(--brand);
  font-size: 1.3em;
  font-weight: 900;
}

.progress-big {
  margin-top: 0.35em;
  color: var(--text-strong);
  font-size: 4.5em;
  font-weight: 900;
  line-height: 1;
}

.progress-sub {
  margin-top: 0.5em;
  color: var(--text-muted);
  font-weight: 700;
}

.bar {
  height: 0.75em;
  margin: 1.5em 0;
  overflow: hidden;
  border-radius: 99px;
  background: #E2E8F0;
}

.bar span {
  display: block;
  height: 100%;
  border-radius: inherit;
  background: linear-gradient(90deg, var(--brand), var(--success));
}

.progress-stats {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 0.75em;
}

.progress-stats div {
  background: #F1F5F9;
  border-radius: 7px;
  padding: 1em;
}

.progress-stats b {
  display: block;
  color: var(--text-strong);
  font-size: 1.4em;
}

.progress-stats span {
  display: block;
  margin-top: 0.25em;
  color: var(--text-muted);
  font-size: 0.8em;
  font-weight: 700;
}

.reward-overview .hero-metrics {
  gap: 1.25em;
  margin-bottom: 1.9em;
}

.reward-overview .hero-metrics div {
  padding: 1.85em;
}

.reward-overview .hero-metrics span {
  font-size: 3.55em;
}

.reward-overview .progress-cards {
  gap: 1.55em;
}

.reward-overview .progress-card {
  padding: 2.2em;
}

.reward-overview .progress-title {
  font-size: 1.65em;
}

.reward-overview .progress-big {
  font-size: 5.25em;
}

.reward-overview .progress-sub {
  font-size: 1.16em;
}

.reward-overview .progress-stats {
  gap: 1em;
}

.reward-overview .progress-stats b {
  font-size: 1.85em;
}

.reward-overview .progress-stats span {
  font-size: 0.95em;
}

.progress-columns {
  align-items: start;
}

.mini-table {
  width: 100%;
  border-collapse: collapse;
  table-layout: fixed;
  overflow: hidden;
  border-radius: 8px;
  background: #fff;
  border: 1px solid #DDE6EF;
  font-size: 1.02em;
}

.mini-table th {
  background: #102332;
  color: #fff;
  padding: 0.78em;
  text-align: right;
  font-size: 0.8em;
}

.mini-table th:nth-child(1),
.mini-table th:nth-child(2) {
  text-align: left;
}

.mini-table td {
  border-top: 1px solid #E7EDF3;
  padding: 0.74em 0.78em;
  text-align: right;
  white-space: nowrap;
}

.mini-table td:nth-child(1) {
  width: 3em;
  text-align: left;
  color: var(--text-muted);
}

.mini-table td:nth-child(2),
.mini-table th:nth-child(2) {
  width: 12.5em;
  text-align: left;
}

.mini-table td:nth-child(5),
.mini-table th:nth-child(5) {
  width: 10em;
}

.mini-table strong {
  display: inline;
  color: var(--text-strong);
}

.mini-table .person-unit {
  display: inline;
  margin-left: 0.35em;
  color: var(--text-muted);
  font-size: 0.84em;
  font-weight: 700;
}

.status-badge {
  display: inline-flex;
  justify-content: center;
  min-width: 6.2em;
  border-radius: 999px;
  padding: 0.32em 0.8em;
  font-size: 0.92em;
  font-weight: 900;
}

.status-badge.ok {
  background: rgba(48,218,162,0.14);
  color: #128260;
}

.status-badge.wait {
  background: rgba(255,137,51,0.13);
  color: #A24D00;
}

.table-scroll {
  overflow: auto;
  max-height: 780px;
  border-radius: 8px;
  box-shadow: inset 0 0 0 1px #DDE6EF;
}

.deck .slide:not(.cover) {
  position: relative;
  overflow: hidden;
  background:
    linear-gradient(180deg, rgba(255,255,255,0.96), rgba(246,249,252,0.98)),
    linear-gradient(90deg, rgba(0,147,193,0.08), rgba(255,201,79,0.08));
  color: var(--text);
}

.deck .slide:not(.cover)::before {
  content: "";
  position: absolute;
  top: 0;
  left: 0;
  right: 0;
  height: 0.55em;
  background: linear-gradient(90deg, var(--brand), var(--success), var(--gold), var(--orange));
}

.deck .slide:not(.cover)::after {
  content: "";
  position: absolute;
  left: 3.35em;
  right: 3.35em;
  bottom: 3.2em;
  height: 1px;
  background: linear-gradient(90deg, rgba(0,147,193,0.38), rgba(255,201,79,0.28), transparent);
  pointer-events: none;
}

.deck .slide-header {
  position: relative;
  z-index: 1;
  align-items: center;
  min-height: 8.15em;
  margin-bottom: 1.05em;
}

.deck .slide-header > div:first-child {
  position: relative;
  padding-left: 1.05em;
}

.deck .slide-header > div:first-child::before {
  content: "";
  position: absolute;
  left: 0;
  top: 0.18em;
  bottom: 0.18em;
  width: 0.22em;
  border-radius: 8px;
  background: linear-gradient(180deg, var(--brand), var(--gold));
}

.deck .slide-header h2 {
  max-width: 15.5em;
  font-size: 3.72em;
  letter-spacing: 0;
}

.deck .meta {
  color: #5E7084;
  font-weight: 900;
  letter-spacing: 0.14em;
}

.deck .pill {
  flex: 0 0 auto;
  border: 1px solid rgba(0,147,193,0.16);
  background: #EEF8FB;
  color: #0E7698;
  box-shadow: 0 0.65em 1.6em rgba(0,147,193,0.08);
  font-size: 1.05em;
  padding: 0.55em 0.9em;
}

.header-actions {
  display: flex;
  align-items: center;
  gap: 0.75em;
  flex: 0 0 auto;
}

.pdf-export-btn {
  border: 1px solid rgba(0,147,193,0.22);
  border-radius: 8px;
  padding: 0.58em 0.92em;
  background: #FFFFFF;
  color: #0E7698;
  box-shadow: 0 0.65em 1.5em rgba(0,147,193,0.08);
  cursor: pointer;
  font: inherit;
  font-size: 1.02em;
  font-weight: 900;
  letter-spacing: 0;
}

.pdf-export-btn:hover {
  background: #EEF8FB;
}

.deck .pill.success {
  border-color: rgba(48,218,162,0.22);
  background: #EAFBF5;
  color: #0D8561;
}

.deck .pill.orange,
.deck .pill.gold {
  border-color: rgba(255,137,51,0.22);
  background: #FFF1E8;
  color: #A05216;
}

.deck .hero-metrics {
  gap: 1.2em;
  margin-bottom: 1.15em;
}

.deck .hero-metrics div,
.deck .summary-strip div,
.deck .progress-card,
.deck .rule-card {
  position: relative;
  overflow: hidden;
  background: #FFFFFF;
  color: var(--text-strong);
  border: 1px solid rgba(148,163,184,0.32);
  border-radius: 8px;
  box-shadow: 0 1.1em 2.6em rgba(15,23,42,0.075);
}

.deck .hero-metrics div::before,
.deck .summary-strip div::before,
.deck .progress-card::before {
  content: "";
  position: absolute;
  top: 0;
  left: 0;
  right: 0;
  height: 0.28em;
  background: var(--brand);
}

.deck .hero-metrics div:nth-child(2)::before,
.deck .summary-strip div:nth-child(2)::before,
.deck .progress-card:nth-child(2)::before {
  background: var(--success);
}

.deck .hero-metrics div:nth-child(3)::before,
.deck .summary-strip div:nth-child(3)::before {
  background: var(--gold);
}

.deck .summary-strip div:nth-child(4)::before {
  background: var(--orange);
}

.deck .hero-metrics span,
.deck .summary-strip span,
.deck .detail-metrics span {
  color: var(--text-strong);
  letter-spacing: 0;
}

.deck .hero-metrics label,
.deck .summary-strip label,
.deck .detail-metrics label {
  color: #5E7084;
  font-weight: 900;
  letter-spacing: 0.04em;
}

.deck .hero-metrics span {
  font-size: 3.38em;
}

.deck .summary-strip {
  gap: 1em;
  margin-bottom: 1.25em;
}

.deck .summary-strip div {
  padding: 1.22em 1.35em;
}

.deck .summary-strip span {
  font-size: 2.55em;
}

.deck .progress-cards {
  gap: 1.2em;
}

.deck .progress-card {
  padding: 1.75em;
}

.deck .progress-title {
  color: #0E7698;
  font-size: 1.42em;
  letter-spacing: 0;
}

.deck .progress-big {
  color: var(--text-strong);
  font-size: 4.85em;
  letter-spacing: 0;
}

.deck .progress-sub {
  color: #5E7084;
  font-size: 1.06em;
}

.deck .bar {
  height: 0.9em;
  background: #E8EEF5;
  box-shadow: inset 0 0.06em 0.22em rgba(15,23,42,0.08);
}

.deck .bar span {
  position: relative;
  background: linear-gradient(90deg, var(--brand), var(--success));
  animation: barGrow 0.72s cubic-bezier(.22,.8,.26,.99) both;
  transform-origin: left center;
}

.deck .rule-grid {
  gap: 1.35em;
}

.deck .rule-card {
  min-height: 31em;
  padding: 2.05em;
}

.deck .rule-card.primary {
  border-top: 0;
}

.deck .rule-card.accent {
  border-top: 0;
}

.deck .rule-card.primary::before,
.deck .rule-card.accent::before {
  content: "";
  position: absolute;
  inset: 0 auto 0 0;
  width: 0.42em;
  background: linear-gradient(180deg, var(--brand), var(--success));
}

.deck .rule-card.accent::before {
  background: linear-gradient(180deg, var(--orange), var(--gold));
}

.deck .rule-card h3 {
  font-size: 2.18em;
  letter-spacing: 0;
}

.deck .rule-card li {
  font-size: 1.24em;
  line-height: 1.52;
}

.deck .rule-amount {
  background: #F6FAFD;
  border: 1px solid #DCE8F2;
  color: var(--text-strong);
  box-shadow: inset 0 0 0 1px rgba(255,255,255,0.72);
}

.deck .rule-note {
  border: 1px solid rgba(255,201,79,0.45);
  border-left: 0.45em solid var(--gold);
  background: #FFF9E8;
  color: #684B00;
  box-shadow: 0 0.8em 2em rgba(104,75,0,0.07);
}

.deck .data-table {
  position: relative;
  z-index: 1;
  overflow: hidden;
  border: 1px solid rgba(148,163,184,0.30);
  background: #FFFFFF;
  box-shadow: 0 1.15em 2.8em rgba(15,23,42,0.065);
}

.deck .data-table th {
  background: #122432;
  border-right: 1px solid rgba(255,255,255,0.08);
  color: #FFFFFF;
  letter-spacing: 0.03em;
}

.deck .data-table td {
  border-top: 1px solid #E8EEF5;
}

.deck .data-table tbody tr:nth-child(even) td {
  background: #FAFCFE;
}

.deck .data-table tbody tr:nth-child(1) td,
.deck .data-table tbody tr:nth-child(2) td,
.deck .data-table tbody tr:nth-child(3) td {
  background: linear-gradient(90deg, rgba(255,201,79,0.15), rgba(255,255,255,0) 48%);
}

.deck .data-table tbody tr:nth-child(1) td:first-child,
.deck .data-table tbody tr:nth-child(2) td:first-child,
.deck .data-table tbody tr:nth-child(3) td:first-child {
  color: #B87300;
  font-weight: 900;
}

.deck .data-table td strong {
  letter-spacing: 0;
}

.deck .data-table .person-unit {
  color: #66778D;
}

.deck .status-badge {
  border: 1px solid transparent;
  box-shadow: inset 0 -1px 0 rgba(255,255,255,0.4);
}

.deck .status-badge.ok {
  border-color: rgba(48,218,162,0.22);
  background: #E8FAF3;
  color: #0D7F5D;
}

.deck .status-badge.wait {
  border-color: rgba(255,137,51,0.22);
  background: #FFF0E6;
  color: #98531B;
}

.deck .focus-slide,
.deck .ranking-slide {
  padding-top: 4.15em;
  padding-bottom: 2.35em;
}

.deck .focus-slide .slide-header,
.deck .ranking-slide .slide-header {
  min-height: 7.1em;
  margin-bottom: 0.72em;
}

.deck .focus-slide .slide-header h2,
.deck .ranking-slide .slide-header h2 {
  font-size: 4.08em;
}

.deck .focus-slide .data-table,
.deck .ranking-slide .data-table {
  font-size: 1.37em;
}

.deck .focus-slide .data-table th,
.deck .ranking-slide .data-table th {
  padding-top: 0.7em;
  padding-bottom: 0.7em;
}

.deck .focus-slide .data-table td,
.deck .ranking-slide .data-table td {
  padding-top: 0.76em;
  padding-bottom: 0.76em;
}

.deck .travel-slide .data-table {
  margin-bottom: 0.68em;
}

.deck .travel-note {
  margin-top: 0.8em;
  font-size: 0.98em;
}

.deck .reward-overview .hero-metrics span {
  font-size: 3.72em;
}

.deck .reward-overview .progress-card {
  padding: 2em;
}

.deck .reward-overview .progress-big {
  font-size: 5.35em;
}

.deck .table-scroll {
  border-radius: 8px;
  box-shadow: 0 1.1em 2.6em rgba(15,23,42,0.06);
}

@keyframes barGrow {
  from { transform: scaleX(0); }
  to { transform: scaleX(1); }
}

@keyframes stageRise {
  from {
    opacity: 0;
    transform: translateY(1.1em);
  }
  to {
    opacity: 1;
    transform: translateY(0);
  }
}

.deck .slide-header,
.deck .hero-metrics > div,
.deck .summary-strip > div,
.deck .progress-card,
.deck .rule-card,
.deck .rule-note,
.deck .data-table {
  animation: stageRise 0.48s cubic-bezier(.2,.8,.24,1) both;
}

.deck .hero-metrics > div:nth-child(1),
.deck .summary-strip > div:nth-child(1),
.deck .progress-card:nth-child(1),
.deck .rule-card:nth-child(1) {
  animation-delay: 0.06s;
}

.deck .hero-metrics > div:nth-child(2),
.deck .summary-strip > div:nth-child(2),
.deck .progress-card:nth-child(2),
.deck .rule-card:nth-child(2) {
  animation-delay: 0.12s;
}

.deck .hero-metrics > div:nth-child(3),
.deck .summary-strip > div:nth-child(3),
.deck .rule-note {
  animation-delay: 0.18s;
}

.deck .data-table {
  animation-delay: 0.16s;
}

@media (prefers-reduced-motion: reduce) {
  .deck .slide-header,
  .deck .hero-metrics > div,
  .deck .summary-strip > div,
  .deck .progress-card,
  .deck .rule-card,
  .deck .rule-note,
  .deck .data-table,
  .deck .bar span {
    animation: none;
  }
}

.full-table {
  font-size: 0.82em;
  border: 0;
}

.full-table th {
  font-size: 0.76em;
}

.full-table td {
  padding-top: 0.52em;
  padding-bottom: 0.52em;
}

.full-table td:nth-child(2),
.full-table th:nth-child(2) {
  width: 15em;
}

.full-table thead th {
  position: sticky;
  top: 0;
  z-index: 2;
}

.empty {
  display: grid;
  place-items: center;
  min-height: 18em;
  color: var(--text-muted);
  background: #F1F5F9;
  border: 1px dashed #CBD5E1;
  border-radius: 8px;
}

html:not(.grid-mode),
html:not(.grid-mode) body {
  overflow: hidden;
  overscroll-behavior: none;
}

.ctrl-bar button[title="投影片 ↔ 捲動模式"] {
  display: none;
}

.pdf-report-root {
  display: none;
}

@media print {
  @page {
    size: A4 portrait;
    margin: 10mm;
  }

  body.pdf-exporting {
    background: #FFFFFF !important;
  }

  body.pdf-exporting .deck,
  body.pdf-exporting .ctrl-bar,
  body.pdf-exporting .lightbox {
    display: none !important;
  }

  body.pdf-exporting .pdf-report-root {
    display: block !important;
  }

  body.pdf-exporting .pdf-report {
    color: #17212B;
    font-family: "Noto Sans TC", "PingFang TC", "Microsoft JhengHei", Arial, sans-serif;
    font-size: 8pt;
    line-height: 1.25;
  }

  body.pdf-exporting .pdf-report-header {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    gap: 8mm;
    border-bottom: 2px solid #122432;
    padding-bottom: 4mm;
    margin-bottom: 4mm;
  }

  body.pdf-exporting .pdf-report-header h1 {
    margin: 0 0 1.5mm;
    color: #0B1720;
    font-size: 16pt;
    line-height: 1.15;
  }

  body.pdf-exporting .pdf-report-meta {
    color: #5E7084;
    font-size: 8pt;
    font-weight: 700;
  }

  body.pdf-exporting .pdf-report-summary {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 2mm;
    margin-bottom: 4mm;
  }

  body.pdf-exporting .pdf-report-summary div {
    border: 1px solid #D8E1EA;
    border-radius: 2mm;
    padding: 2mm;
  }

  body.pdf-exporting .pdf-report-summary b {
    display: block;
    color: #0B1720;
    font-size: 11pt;
    line-height: 1.1;
  }

  body.pdf-exporting .pdf-report-summary span {
    display: block;
    margin-top: 1mm;
    color: #5E7084;
    font-size: 7pt;
    font-weight: 700;
  }

  body.pdf-exporting .pdf-report-table {
    width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
  }

  body.pdf-exporting .pdf-report-table thead {
    display: table-header-group;
  }

  body.pdf-exporting .pdf-report-table th {
    background: #122432 !important;
    color: #FFFFFF !important;
    font-size: 7pt;
    font-weight: 900;
    padding: 1.7mm 1.2mm;
    text-align: right;
  }

  body.pdf-exporting .pdf-report-table td {
    border-bottom: 1px solid #E1E8F0;
    color: #17212B;
    font-size: 7.1pt;
    padding: 1.35mm 1.2mm;
    text-align: right;
    vertical-align: middle;
    white-space: nowrap;
  }

  body.pdf-exporting .pdf-report-table th:nth-child(1),
  body.pdf-exporting .pdf-report-table td:nth-child(1),
  body.pdf-exporting .pdf-report-table th:nth-child(2),
  body.pdf-exporting .pdf-report-table td:nth-child(2),
  body.pdf-exporting .pdf-report-table th:nth-child(3),
  body.pdf-exporting .pdf-report-table td:nth-child(3),
  body.pdf-exporting .pdf-report-table th:nth-child(9),
  body.pdf-exporting .pdf-report-table td:nth-child(9) {
    text-align: left;
  }

  body.pdf-exporting .pdf-report-table th:nth-child(1),
  body.pdf-exporting .pdf-report-table td:nth-child(1) {
    width: 6mm;
  }

  body.pdf-exporting .pdf-report-table th:nth-child(2),
  body.pdf-exporting .pdf-report-table td:nth-child(2) {
    width: 18mm;
  }

  body.pdf-exporting .pdf-report-table th:nth-child(3),
  body.pdf-exporting .pdf-report-table td:nth-child(3) {
    width: 19mm;
  }

  body.pdf-exporting .pdf-report-table th:nth-child(9),
  body.pdf-exporting .pdf-report-table td:nth-child(9) {
    width: 18mm;
  }

  body.pdf-exporting .pdf-report-table tr {
    break-inside: avoid;
    page-break-inside: avoid;
  }
}
"""


def build() -> None:
    records = extract_records()
    summary = summarize(records)
    total_pages = 11
    slides = [
        cover(summary, total_pages),
        overview(summary, total_pages),
        rules_slide(total_pages),
        travel_slide(records, summary, total_pages),
        reward_summary_slide(summary, total_pages, page=5),
        reach_progress_slide(records, total_pages, page=6),
        over_progress_slide(records, total_pages, page=7),
        ranking_slide(records, "個人組", 8, total_pages),
        ranking_slide(records, "新人組", 9, total_pages),
        full_data_slide(records, "個人組", 10, total_pages),
        full_data_slide(records, "新人組", 11, total_pages),
    ]

    template = TEMPLATE.read_text(encoding="utf-8")
    template = template.replace("<title>Deck Title</title>", f"<title>{REPORT_TITLE_FULL}</title>")
    template = re.sub(
        r"function toggleMode\(\) \{\n  mode = mode === 'slide' \? 'scroll' : 'slide';\n  document\.body\.classList\.toggle\('scroll-mode'\);\n  document\.documentElement\.classList\.toggle\('scroll-mode'\);\n  render\(\);\n\}",
        "function toggleMode() {\n"
        "  mode = 'slide';\n"
        "  document.body.classList.remove('scroll-mode');\n"
        "  document.documentElement.classList.remove('scroll-mode');\n"
        "  render();\n"
        "}",
        template,
        count=1,
    )
    template = re.sub(
        r"/\* ── 手機：自動切換捲動模式 ── \*/\nif \(window\.innerWidth <= 768\) \{\n  mode = 'scroll';\n  document\.body\.classList\.add\('scroll-mode'\);\n  document\.documentElement\.classList\.add\('scroll-mode'\);\n\}",
        "/* ── 投影模式：不自動切換成捲動長頁 ── */",
        template,
        count=1,
    )
    template = re.sub(r'<link rel="stylesheet" href="../css/variables.css">\n', "", template)
    template = template.replace("/* ══════════════════════════════════════════════════════\n   每張 deck 的個別覆蓋 — 只在此處覆蓋，不改 variables.css\n   大多數情況下留空即可\n   ══════════════════════════════════════════════════════ */", custom_css())
    template = re.sub(
        r"<div class=\"deck\">\s*<!-- 封面範例 -->.*?</div>\n\n<!-- 控制列 -->",
        '<div class="deck">\n' + "\n".join(slides) + "\n</div>\n\n<!-- 控制列 -->",
        template,
        flags=re.S,
        count=1,
    )
    payload = {
        "source": str(WORKBOOK),
        "targets": TARGET_OFFICES,
        "groups": GROUPS,
        "count": len(records),
        "summary": summary,
        "records": records,
    }
    xlsx_lib = XLSX_LIB.read_text(encoding="utf-8").replace("</script", "<\\/script")
    excel_refresh = EXCEL_REFRESH.read_text(encoding="utf-8").replace("</script", "<\\/script")
    template = template.replace(
        "</body>",
        f"<script>{xlsx_lib}</script>\n"
        f'<script id="excel-data" type="application/json">{json.dumps(payload, ensure_ascii=False)}</script>\n'
        f"<script>{excel_refresh}</script>\n"
        "</body>",
    )
    OUTPUT.write_text(template, encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"Records: {len(records)}")
    for group in GROUPS:
        print(group, summary[group]["count"], summary[group]["contest_total"])


if __name__ == "__main__":
    build()
